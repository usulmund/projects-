
import sys
from pdf2docx import Converter
from docx.api import Document
import pandas as pd
import openpyxl
import os
from tkinter import *
from tkinter.ttk import *
from tkinter import font
import re

def clicked():
    pdf = txt.get()
    docx = pdf.split(".")[0] + ".docx"
    xlsx = pdf.split(".")[0] + ".xlsx"


    if (os.path.exists(pdf)):
        if (os.path.exists(xlsx) == False):
            cv = Converter(pdf)
            cv.convert(docx)

            document = Document(docx)
            df = pd.DataFrame()

            for table in document.tables:
                for row in table.rows:
                    text = [cell.text for cell in row.cells]
                    df = df.append([text], ignore_index=True)

            df.to_excel(xlsx)

        from openpyxl import load_workbook
        import datetime

        wb = load_workbook(xlsx)
        sheet = wb.active
        i = 4
        lesson = ""
        aud = ""
        less = 0
        aud_1610 = 0
        aud_1617 = 0
        aud_1610_h = 0
        aud_1617_h = 0

        teacher = sheet.cell(row=2, column=3).value
        types = {}
        types_h = {}

        while lesson != None and aud != None:
            lesson = sheet.cell(row=i, column=3).value
            aud = sheet.cell(row=i, column=6).value
            lesson_name = sheet.cell(row=i, column=4).value

            if lesson != "Занятий нет" and lesson != None and aud != None and 'Отменено' not in lesson_name:
                les_type = sheet.cell(row=i, column=7).value
                types[les_type] = types.get(les_type, 0) + 1
                less += 1
                l = lesson.split(" - ")
                start = datetime.datetime.strptime(l[0], "%H:%M")
                end = datetime.datetime.strptime(l[1], "%H:%M")
                delta_h = end - start
                akad_l = list(map(int, str(delta_h).split(":")))
                acad_m = akad_l[0] * 60 + akad_l[1]

                types_h[les_type] = types_h.get(les_type, 0) + acad_m

                if aud == "1610":
                    aud_1610 += 1
                    aud_1610_h += acad_m
                if aud == "1617":
                    aud_1617 += 1
                    aud_1617_h += acad_m

            i += 1

        C = types.get('С',0)
        C_h = types_h.get('С',0)

        L = types.get('Л',0)
        L_h = types_h.get('Л',0)

        LR = types.get('ЛР',0)
        LR_h = types_h.get('ЛР',0)

        KE = types.get('КЭ',0)
        KE_h = types_h.get('КЭ',0)

        E = types.get('Э',0)
        E_h = types_h.get('Э',0)

        M = types.get('М',0)
        M_h = types_h.get('М',0)

        res = openpyxl.load_workbook('Фактическая нагрузка.xlsx')
        worksheet = res['Нагрузка преподавателей']
        find_line = 5
        cell_for_teacher = worksheet.cell(row=find_line, column=2).value
        while (cell_for_teacher != None and cell_for_teacher != teacher):
            find_line += 1
            cell_for_teacher = worksheet.cell(row=find_line, column=2).value

        worksheet['B' + str(find_line)] = teacher
        worksheet['F' + str(find_line)] = str(round(L_h/45.0, 2))
        worksheet['G' + str(find_line)] = str(round(C_h / 45.0, 2))
        worksheet['H' + str(find_line)] = str(round(LR_h / 45.0, 2))
        worksheet['I' + str(find_line)] = str(round(E_h / 45.0, 2))
        worksheet['K' + str(find_line)] = sum(list(map(float, [worksheet.cell(row=find_line, column=fj).value for fj in range(6, 11) if worksheet.cell(row=find_line, column=fj).value != None])))#F - J

        worksheet['T' + str(find_line)] = str(round(KE_h / 45.0, 2))
        worksheet['AC' + str(find_line)] = str(round(M_h / 45.0, 2))
        worksheet['AE' + str(find_line)] = sum(list(map(float, [worksheet.cell(row=find_line, column=lad).value for lad in range(12, 31) if worksheet.cell(row=find_line, column=lad).value != None])))#L - AD

        worksheet['AF' + str(find_line)] = str(round(float(worksheet.cell(row=find_line, column=11).value) + float(worksheet.cell(row=find_line, column=31).value), 2))

        worksheet = res['Загрузка 1610 и 1617']
        find_line = 4
        cell_for_teacher = worksheet.cell(row=find_line, column=1).value
        while (cell_for_teacher != None and cell_for_teacher != teacher):
            find_line += 1
            cell_for_teacher = worksheet.cell(row=find_line, column=1).value

        worksheet['A' + str(find_line)] = teacher
        worksheet['B' + str(find_line)] = str(round(aud_1610_h / 45.0, 2))
        worksheet['C' + str(find_line)] = str(round(aud_1617_h / 45.0, 2))

        res.save('Фактическая нагрузка.xlsx')

        lbl.configure(text = "\n" + teacher + "\n\nКоличeство академических часов преподавателя (без учёта экзаменов): " + str(round((C_h + L_h + LR_h + KE_h + M_h)/ 45.0, 2)) +
                           "\nКоличeство пар: " + str(C + L + LR + KE + M) +
                           "\n\nКоличeство семинаров: " + str(C) +
                           "\nКоличeство академических часов на семинары: " + str(round(C_h/45.0, 2)) +
                           "\n\nКоличeство лекций: " + str(L) +
                           "\nКоличeство академических часов на лекции: " + str(round(L_h/ 45.0, 2)) +
                           "\n\nКоличeство лабораторных работ: " + str(LR) +
                           "\nКоличeство академических часов на лабораторные работы: " + str(round(LR_h/ 45.0, 2)) +
                           "\n\nКоличeство консультаций к экзамену: " + str(KE) +
                           "\nКоличeство академических часов на консультации к экзамену: " + str(round(KE_h/ 45.0, 2)) +
                           "\n\nКоличeство экзаменов: " + str(E) +
                           "\nКоличeство академических часов на экзамены : " + str(round(E_h/ 45.0, 2)) +
                           "\n\nКоличeство пар на прием текущих задолженностей: " + str(M) +
                           "\nКоличeство академических часов на прием текущих задолженностей : " + str(round(M_h / 45.0, 2)) +
                           "\n\nЗагруженность 1610 в академических часах: " + str(round(aud_1610_h / 45.0, 2)) +
                           "\nЗагруженность 1610 по количеству пар: " + str(aud_1610) +
                           "\nЗагруженность 1617 в академических часах: " + str(round(aud_1617_h / 45.0, 2)) +
                           "\nЗагруженность 1617 по количеству пар: " + str(aud_1617))
    else:
        lbl.configure(text="Такой pdf-файл отсутствует")



window = Tk()
window["bg"] = "white"

window.title("ИНФОРМАЦИЯ")
window.geometry('890x550')


font1 = font.Font(family="Open Sans", size = 12)
lbl = Label(window,
            font=font1,
            text="Введите название pdf-файла: ",
            foreground="black",
            background="white",
            )
lbl.grid(column=0, row=0)

txt = Entry(window,
            font=font1,
            width=20,
            background="white"
            )
txt.grid(column=1, row=0)

btn = Button(window,
             text="ОБРАБОТАТЬ",
             command=clicked)
btn.grid(column=2, row=0)
window.mainloop()


